#!/usr/bin/env python3
"""
StoreLense — ERP SOH CSV Generator
====================================
Reads the storelense_products.xlsx (output of convert_to_storelense.py) and
produces an ERP SOH CSV ready to be ingested by the ERP integration service.

The ERP ingestion service (ErpCsvParser.java) requires exactly four columns:
  EAN            — EAN-13 barcode (same for all units of a product)
  EXPECTED_QTY   — integer quantity ERP expects in the store
  ZONE_REGION    — zone code (e.g. SALES_FLOOR, BACKROOM). Leave blank for store-total.
  STORE_CODE     — ERP store code (e.g. EW001); must exist in erp.erp_store_mapping table.

Usage:
  python tools/generate_erp_soh.py --input storelense_products.xlsx --store-code EW001
  python tools/generate_erp_soh.py --input storelense_products.xlsx --store-code EW001 --zone SALES_FLOOR
  python tools/generate_erp_soh.py --input storelense_products.xlsx --store-code EW001 --split-zones
  python tools/generate_erp_soh.py --sample --store-code EW001   # create a sample CSV without reading xlsx

Upload to server:
  scp erp_soh_YYYY-MM-DD.csv root@SERVER_IP:/tmp/erp-import/

Or trigger via Admin API (auto-picked up from local folder if localEnabled=true):
  curl -s -X POST http://SERVER_IP:8080/api/erp/admin/import \\
       -H "Authorization: Bearer <admin_token>" \\
       -H "Content-Type: application/json" \\
       -d '{"path":"/tmp/erp-import/erp_soh_YYYY-MM-DD.csv","type":"FILE"}'
"""

import argparse
import csv
import os
import sys
from datetime import date

try:
    import openpyxl
except ImportError:
    sys.exit("Run: pip install openpyxl")

GREEN  = "\033[92m"
RED    = "\033[91m"
YELLOW = "\033[93m"
CYAN   = "\033[96m"
BOLD   = "\033[1m"
RESET  = "\033[0m"


SAMPLE_DATA = [
    # EAN,           EXPECTED_QTY, ZONE_REGION
    ("8905454700001", 4,  "SALES_FLOOR"),
    ("8905454700001", 2,  "BACKROOM"),
    ("8905454700002", 3,  "SALES_FLOOR"),
    ("8905454700003", 5,  "SALES_FLOOR"),
    ("8905454700004", 2,  "SALES_FLOOR"),
    ("8905454700005", 6,  "SALES_FLOOR"),
    ("8905454700005", 1,  "BACKROOM"),
]


def write_csv(rows, output_path: str, store_code: str):
    """Write ERP SOH CSV. rows = list of (ean, qty, zone)."""
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["EAN", "EXPECTED_QTY", "ZONE_REGION", "STORE_CODE"])
        for ean, qty, zone in rows:
            writer.writerow([ean, qty, zone or "", store_code])
    print(f"\n  {GREEN}[OK]{RESET}   ERP SOH CSV saved → {BOLD}{output_path}{RESET}")
    print(f"         {len(rows)} row(s) written  |  store: {BOLD}{store_code}{RESET}")


def create_sample(output_path: str, store_code: str):
    write_csv(SAMPLE_DATA, output_path, store_code)
    print(f"\n  This is a sample. Edit EAN values to match your actual products.")
    print(f"  EAN values must match the 'barcode' column in storelense_products.xlsx")
    print(f"  (or the EAN-13 printed on your ERP items).\n")


def read_xlsx(xlsx_path: str, zone_override: str | None, split_zones: bool) -> list:
    """
    Read storelense_products.xlsx produced by convert_to_storelense.py.
    Groups rows by EAN, sums expected qty.
    """
    print(f"  Reading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active

    iter_rows = ws.iter_rows(values_only=True)
    raw_hdr = next(iter_rows)
    headers = [str(h).strip().lower() if h else f"col{i}" for i, h in enumerate(raw_hdr)]

    # Locate columns — accept both 'barcode' and 'ean' names
    def col_idx(*names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return None

    ean_col      = col_idx("barcode", "ean", "ean13")
    style_col    = col_idx("style", "sku")
    expected_col = col_idx("expected")
    floor_col    = col_idx("sales floor", "salesfloor")
    back_col     = col_idx("back stock", "backstock")

    if ean_col is None:
        sys.exit(f"{RED}Column 'barcode' or 'ean' not found in {xlsx_path}{RESET}")
    if expected_col is None:
        sys.exit(f"{RED}Column 'expected' not found in {xlsx_path}{RESET}")

    # Aggregate per EAN
    by_ean: dict[str, dict] = {}  # ean → {total, floor, back}

    for row in iter_rows:
        if not any(v is not None for v in row):
            continue
        ean = str(row[ean_col] or "").strip()
        if not ean or not ean.isdigit():
            continue

        try:
            exp   = int(row[expected_col] or 0)
            floor = int(row[floor_col] or 0) if floor_col is not None else exp
            back  = int(row[back_col]  or 0) if back_col  is not None else 0
        except (ValueError, TypeError):
            continue

        if ean not in by_ean:
            by_ean[ean] = {"total": 0, "floor": 0, "back": 0}
        by_ean[ean]["total"] += exp
        by_ean[ean]["floor"] += floor
        by_ean[ean]["back"]  += back

    wb.close()
    print(f"  Products (unique EAN): {len(by_ean)}")

    rows = []
    if split_zones:
        # One row per EAN per zone that has qty > 0
        for ean, d in by_ean.items():
            if d["floor"] > 0:
                rows.append((ean, d["floor"], "SALES_FLOOR"))
            if d["back"] > 0:
                rows.append((ean, d["back"],  "BACKROOM"))
            if d["floor"] == 0 and d["back"] == 0 and d["total"] > 0:
                rows.append((ean, d["total"], ""))
    else:
        zone = zone_override or ""
        for ean, d in by_ean.items():
            qty = d["total"]
            if qty > 0:
                rows.append((ean, qty, zone))

    return rows


def main():
    ap = argparse.ArgumentParser(
        description="Generate ERP SOH CSV for StoreLense ingestion service",
        epilog="""
Examples:
  python tools/generate_erp_soh.py --sample
  python tools/generate_erp_soh.py --input storelense_products.xlsx
  python tools/generate_erp_soh.py --input storelense_products.xlsx --zone SALES_FLOOR
  python tools/generate_erp_soh.py --input storelense_products.xlsx --split-zones
        """,
        formatter_class=argparse.RawDescriptionHelpFormatter)

    ap.add_argument("--input",       metavar="FILE",  help="storelense_products.xlsx (from convert_to_storelense.py)")
    ap.add_argument("--output",      metavar="FILE",  help="Output CSV filename (default: erp_soh_YYYY-MM-DD.csv)")
    ap.add_argument("--store-code",  metavar="CODE",  help="ERP store code (e.g. EW001) — must exist in erp_store_mapping", required=False, default="EW001")
    ap.add_argument("--zone",        metavar="ZONE",  help="Zone code for all rows (e.g. SALES_FLOOR, BACKROOM). Leave blank for store-total.")
    ap.add_argument("--split-zones", action="store_true", help="Write separate rows for SALES_FLOOR and BACKROOM using 'sales floor'/'back stock' columns")
    ap.add_argument("--sample",      action="store_true", help="Generate a sample CSV and exit")

    args = ap.parse_args()
    today = date.today().isoformat()
    default_out = args.output or f"erp_soh_{today}.csv"

    print(f"\n{BOLD}{CYAN}StoreLense ERP SOH CSV Generator{RESET}")

    store_code = args.store_code

    if args.sample:
        create_sample(default_out, store_code)
        _print_upload_instructions(default_out)
        return

    if not args.input:
        ap.print_help()
        print(f"\n{YELLOW}Tip:{RESET} run --sample to create an example CSV first.\n")
        return

    if not os.path.exists(args.input):
        sys.exit(f"{RED}File not found: {args.input}{RESET}")

    rows = read_xlsx(args.input, args.zone, args.split_zones)
    if not rows:
        sys.exit(f"{RED}No valid rows found in {args.input}{RESET}")

    write_csv(rows, default_out, store_code)
    _print_upload_instructions(default_out)


def _print_upload_instructions(csv_file: str):
    print(f"""
{BOLD}Next steps — upload to StoreLense:{RESET}

  {CYAN}Option A — drop in watched folder (auto-ingested):{RESET}
    scp {csv_file} root@SERVER_IP:/tmp/erp-import/
    (service picks it up automatically if localEnabled=true in application.yml)

  {CYAN}Option B — trigger via Admin API:{RESET}
    # 1. Get admin token
    TOKEN=$(curl -s -X POST http://SERVER_IP:8080/api/auth/login \\
      -H 'Content-Type: application/json' \\
      -d '{{"username":"admin","password":"Admin@StoreLense1"}}' | python3 -c "import sys,json; print(json.load(sys.stdin)['data']['accessToken'])")

    # 2. Copy file to server first
    scp {csv_file} root@SERVER_IP:/tmp/erp-import/

    # 3. Trigger import
    curl -s -X POST http://SERVER_IP:8080/api/erp/admin/import \\
      -H "Authorization: Bearer $TOKEN" \\
      -H "Content-Type: application/json" \\
      -d '{{"path":"/tmp/erp-import/{csv_file}","type":"FILE"}}'

  {CYAN}Option C — generate SQL and import directly:{RESET}
    python tools/upload_erp_data.py --inventory {csv_file} --export-sql erp_soh.sql
    scp erp_soh.sql root@SERVER_IP:/tmp/
    ssh root@SERVER_IP "psql -h localhost -U storelense_app -d storelense -f /tmp/erp_soh.sql"

  {CYAN}Check import batches:{RESET}
    curl http://SERVER_IP:8080/api/erp/admin/import/batches \\
      -H "Authorization: Bearer $TOKEN"
""")


if __name__ == "__main__":
    main()
