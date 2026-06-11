#!/usr/bin/env python3
"""Convert product_upload_template.csv → product_upload_template.xlsx"""
import csv, sys
try:
    import openpyxl
except ImportError:
    print("Run: pip install openpyxl"); sys.exit(1)

csv_path  = "tools/product_upload_template.csv"
xlsx_path = "tools/product_upload_template.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Products"

with open(csv_path, newline='', encoding='utf-8') as f:
    for row in csv.reader(f):
        ws.append(row)

# Bold header row
from openpyxl.styles import Font, PatternFill, Alignment
header_fill = PatternFill("solid", fgColor="0F766E")
for cell in ws[1]:
    cell.font      = Font(bold=True, color="FFFFFF")
    cell.fill      = header_fill
    cell.alignment = Alignment(horizontal="center")

# Auto-width columns
for col in ws.columns:
    max_len = max(len(str(c.value or "")) for c in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

wb.save(xlsx_path)
print(f"Saved: {xlsx_path}")
