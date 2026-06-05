#!/usr/bin/env python3
"""
StoreLense — ERP Data Upload Tool
==================================
Uploads ERP SOH (expected inventory) and Product List from XLS/XLSX files
into StoreLense so that RFID scan data can be compared against ERP quantities.

Usage:
  python upload_erp_data.py --create-templates         # Generate sample XLS files
  python upload_erp_data.py --products products.xlsx   # Upload product list
  python upload_erp_data.py --inventory inventory.xlsx # Upload expected inventory (SOH)
  python upload_erp_data.py --products p.xlsx --inventory i.xlsx  # Both at once

Requirements:
  pip install requests openpyxl

Authentication:
  Set STORELENSE_URL, STORELENSE_USER, STORELENSE_PASS as env vars,
  or pass them via --url / --username / --password flags.
"""

import argparse
import json
import os
import sys
import time
from datetime import date, datetime

try:
    import requests
except ImportError:
    sys.exit("Missing dependency: pip install requests openpyxl")

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency: pip install requests openpyxl")

# ── ANSI colours ──────────────────────────────────────────────────────────────
GREEN  = "\033[92m"
RED    = "\033[91m"
YELLOW = "\033[93m"
CYAN   = "\033[96m"
BOLD   = "\033[1m"
RESET  = "\033[0m"

def ok(msg):    print(f"  {GREEN}[OK]{RESET}   {msg}")
def err(msg):   print(f"  {RED}[ERR]{RESET}  {msg}")
def warn(msg):  print(f"  {YELLOW}[WARN]{RESET} {msg}")
def info(msg):  print(f"  {CYAN}[INFO]{RESET} {msg}")
def step(msg):  print(f"\n{BOLD}{CYAN}▶ {msg}{RESET}")

# =============================================================================
# XLS TEMPLATE DEFINITIONS
# =============================================================================

# ── Column definitions ────────────────────────────────────────────────────────

PRODUCT_COLUMNS = [
    # (header,         width, required, notes)
    ("product_code",   18,    True,  "ERP product code — must match ERP system (e.g. ERP-P-00123)"),
    ("sku",            18,    True,  "Stock-keeping unit code — must be unique (e.g. APP-DNM-001)"),
    ("name",           35,    True,  "Product display name (e.g. Classic Denim Jeans)"),
    ("description",    40,    False, "Optional product description"),
    ("category",       18,    False, "Category code: APPAREL, FOOTWEAR, ACCESSORIES, HOMEWARES, ELECTRONICS"),
    ("brand",          18,    False, "Brand name (e.g. StoreLense Basic)"),
    ("uom",            10,    False, "Unit of measure: EACH, PAIR, BOX, ROLL (default: EACH)"),
    ("weight_grams",   14,    False, "Weight in grams — integer (e.g. 650)"),
    ("rfid_enabled",   13,    True,  "TRUE or FALSE — whether this product uses RFID tags"),
    ("epc_tags",       60,    False, "Comma-separated EPC hex values for this product (e.g. 3034257BF400B71400000001,3034257BF400B71400000002)"),
]

INVENTORY_COLUMNS = [
    # (header,         width, required, notes)
    ("store_code",     15,    True,  "Store code matching StoreLense (e.g. SYD001)"),
    ("product_code",   18,    True,  "ERP product code — must match products.product_code"),
    ("sku",            18,    True,  "Product SKU — must match products.sku"),
    ("qty_expected",   14,    True,  "Expected quantity the ERP says should be in this store — integer"),
    ("zone_code",      15,    False, "Zone code (e.g. FLOOR-A, BACKROOM-1). Leave blank for store-level total"),
    ("valid_date",     14,    False, "Date this data is valid for — YYYY-MM-DD (default: today)"),
    ("notes",          30,    False, "Optional notes (e.g. 'post-receiving adjustment')"),
]

PRODUCT_SAMPLE_DATA = [
    ("ERP-P-00001", "APP-DNM-001", "Classic Denim Jeans",       "Regular fit blue denim",    "APPAREL",     "StoreLense Basic", "EACH", 650,  "TRUE",  "3034257BF400B71400000001,3034257BF400B71400000002"),
    ("ERP-P-00002", "APP-TEE-001", "Cotton Crew Tee White",      "100% cotton crew neck",     "APPAREL",     "StoreLense Basic", "EACH", 200,  "TRUE",  "3034257BF400B71400000010,3034257BF400B71400000011"),
    ("ERP-P-00003", "APP-TEE-002", "Cotton Crew Tee Black",      "100% cotton crew neck",     "APPAREL",     "StoreLense Basic", "EACH", 200,  "TRUE",  "3034257BF400B71400000020"),
    ("ERP-P-00004", "FTW-SNK-001", "Classic White Sneaker",      "Lace-up canvas sneaker",    "FOOTWEAR",    "StoreLense Basic", "PAIR", 850,  "TRUE",  "3034257BF400B71400000030"),
    ("ERP-P-00005", "FTW-SND-001", "Summer Sandal",              "Leather sandal",            "FOOTWEAR",    "StoreLense Basic", "PAIR", 450,  "TRUE",  ""),
    ("ERP-P-00006", "ACC-BAG-001", "Canvas Tote Bag",            "Reusable canvas bag",       "ACCESSORIES", "StoreLense Basic", "EACH", 300,  "FALSE", ""),
    ("ERP-P-00007", "HMW-TWL-001", "Cotton Bath Towel",          "600gsm cotton towel",       "HOMEWARES",   "StoreLense Basic", "EACH", 550,  "TRUE",  ""),
]

INVENTORY_SAMPLE_DATA = [
    ("SYD001", "ERP-P-00001", "APP-DNM-001", 50,  "FLOOR-A",    str(date.today()), ""),
    ("SYD001", "ERP-P-00001", "APP-DNM-001", 20,  "BACKROOM-1", str(date.today()), ""),
    ("SYD001", "ERP-P-00002", "APP-TEE-001", 100, "FLOOR-A",    str(date.today()), ""),
    ("SYD001", "ERP-P-00002", "APP-TEE-001", 30,  "BACKROOM-1", str(date.today()), ""),
    ("SYD001", "ERP-P-00003", "APP-TEE-002", 80,  "FLOOR-A",    str(date.today()), ""),
    ("SYD001", "ERP-P-00004", "FTW-SNK-001", 40,  "FLOOR-A",    str(date.today()), "post-delivery adjustment"),
    ("SYD001", "ERP-P-00005", "FTW-SND-001", 25,  "",           str(date.today()), "store-level total"),
    ("SYD001", "ERP-P-00006", "ACC-BAG-001", 200, "FLOOR-A",    str(date.today()), ""),
    ("SYD001", "ERP-P-00007", "HMW-TWL-001", 60,  "BACKROOM-1", str(date.today()), ""),
    ("MEL001", "ERP-P-00001", "APP-DNM-001", 45,  "FLOOR-A",    str(date.today()), ""),
    ("MEL001", "ERP-P-00002", "APP-TEE-001", 90,  "FLOOR-A",    str(date.today()), ""),
]


def _style_header(ws, columns):
    """Apply header row styling: dark background, white bold text."""
    header_fill  = PatternFill("solid", fgColor="1F3864")
    req_fill     = PatternFill("solid", fgColor="2E5FA3")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    for col_idx, (header, width, required, _notes) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font       = header_font
        cell.fill       = req_fill if required else header_fill
        cell.alignment  = center_align
        cell.border     = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _style_data_row(ws, row_idx, col_count):
    """Apply alternating row fill to data rows."""
    fill = PatternFill("solid", fgColor="EBF0FA") if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=False)


def create_product_template(path: str):
    wb = Workbook()

    # ── DATA sheet ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Products"
    _style_header(ws, PRODUCT_COLUMNS)

    for r_idx, row in enumerate(PRODUCT_SAMPLE_DATA, 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
        _style_data_row(ws, r_idx, len(PRODUCT_COLUMNS))

    # ── GUIDE sheet ──────────────────────────────────────────────────────────
    guide = wb.create_sheet("Field Guide")
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 12
    guide.column_dimensions["C"].width = 60
    guide.cell(row=1, column=1, value="Column").font         = Font(bold=True)
    guide.cell(row=1, column=2, value="Required").font       = Font(bold=True)
    guide.cell(row=1, column=3, value="Description").font    = Font(bold=True)

    for r, (col, _w, req, notes) in enumerate(PRODUCT_COLUMNS, 2):
        guide.cell(row=r, column=1, value=col)
        guide.cell(row=r, column=2, value="YES" if req else "no")
        guide.cell(row=r, column=3, value=notes)
        if req:
            guide.cell(row=r, column=2).font = Font(bold=True, color="C00000")

    guide.cell(row=len(PRODUCT_COLUMNS) + 3, column=1,
               value="Blue headers = required fields. Dark headers = optional.").font = Font(italic=True)

    wb.save(path)
    ok(f"Product template saved → {path}")


def create_inventory_template(path: str):
    wb = Workbook()

    ws = wb.active
    ws.title = "ERP_SOH_Expected"
    _style_header(ws, INVENTORY_COLUMNS)

    for r_idx, row in enumerate(INVENTORY_SAMPLE_DATA, 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
        _style_data_row(ws, r_idx, len(INVENTORY_COLUMNS))

    guide = wb.create_sheet("Field Guide")
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 12
    guide.column_dimensions["C"].width = 70
    guide.cell(row=1, column=1, value="Column").font      = Font(bold=True)
    guide.cell(row=1, column=2, value="Required").font    = Font(bold=True)
    guide.cell(row=1, column=3, value="Description").font = Font(bold=True)

    for r, (col, _w, req, notes) in enumerate(INVENTORY_COLUMNS, 2):
        guide.cell(row=r, column=1, value=col)
        guide.cell(row=r, column=2, value="YES" if req else "no")
        guide.cell(row=r, column=3, value=notes)
        if req:
            guide.cell(row=r, column=2).font = Font(bold=True, color="C00000")

    note_row = len(INVENTORY_COLUMNS) + 3
    notes = [
        "• qty_expected: total units the ERP expects in the store (or zone if zone_code is set).",
        "• zone_code: must exactly match a zone code configured in StoreLense (e.g. FLOOR-A).",
        "• Rows with blank zone_code are imported as store-level totals (zone_id = NULL).",
        "• Multiple rows per store+product are allowed (one per zone).",
        "• If a store+product+zone row already exists, it will be UPDATED (upsert).",
        "• valid_date is informational — it is stored in notes but does not affect calculations.",
    ]
    for i, n in enumerate(notes):
        guide.cell(row=note_row + i, column=1, value=n).font = Font(italic=True)

    wb.save(path)
    ok(f"Inventory template saved → {path}")


# =============================================================================
# API CLIENT
# =============================================================================

class StoreLenseClient:
    def __init__(self, base_url: str, username: str, password: str):
        self.base_url    = base_url.rstrip("/")
        self.session     = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        self._token      = None
        self._login(username, password)

    def _login(self, username: str, password: str):
        r = self.session.post(f"{self.base_url}/api/auth/login",
                              json={"username": username, "password": password},
                              timeout=15)
        if r.status_code != 200:
            sys.exit(f"{RED}Login failed ({r.status_code}): {r.text}{RESET}")
        data = r.json().get("data", {})
        self._token = data.get("accessToken")
        self.session.headers["Authorization"] = f"Bearer {self._token}"
        ok(f"Logged in as '{username}' (role: {data.get('role')})")

    def get(self, path: str, **kwargs) -> requests.Response:
        return self.session.get(f"{self.base_url}{path}", timeout=15, **kwargs)

    def post(self, path: str, payload: dict) -> requests.Response:
        return self.session.post(f"{self.base_url}{path}", json=payload, timeout=15)

    def put(self, path: str, payload: dict) -> requests.Response:
        return self.session.put(f"{self.base_url}{path}", json=payload, timeout=15)

    # ── Helpers ──────────────────────────────────────────────────────────────

    def get_stores(self) -> dict:
        """Returns {store_code: store_id}"""
        r = self.get("/api/stores?size=200")
        stores = {}
        if r.status_code == 200:
            for s in r.json().get("data", {}).get("content", []):
                stores[s["storeCode"]] = s["id"]
        return stores

    def get_zones(self, store_id: str) -> dict:
        """Returns {zone_code: zone_id} for a store"""
        r = self.get(f"/api/stores/{store_id}/zones")
        zones = {}
        if r.status_code == 200:
            for z in r.json().get("data", []):
                zones[z["zoneCode"]] = z["id"]
        return zones

    def find_product_by_sku(self, sku: str) -> str | None:
        """Returns product_id or None"""
        r = self.get(f"/api/products/by-sku/{sku}")
        if r.status_code == 200:
            return r.json().get("data", {}).get("id")
        return None


# =============================================================================
# PRODUCT UPLOAD
# =============================================================================

def upload_products(client: StoreLenseClient, xls_path: str,
                    dry_run: bool = False) -> dict:
    step(f"Uploading Products from {xls_path}")

    wb = openpyxl.load_workbook(xls_path, data_only=True)
    ws = wb["Products"]

    headers = [str(ws.cell(1, c).value).strip().lower()
               for c in range(1, ws.max_column + 1)]

    required = {"product_code", "sku", "name", "rfid_enabled"}
    missing  = required - set(headers)
    if missing:
        sys.exit(f"{RED}Missing required columns: {missing}{RESET}")

    stats = {"created": 0, "updated": 0, "skipped": 0, "errors": 0, "epcs": 0}
    col   = {h: i for i, h in enumerate(headers)}

    for row_num in range(2, ws.max_row + 1):
        vals = [ws.cell(row_num, c).value for c in range(1, ws.max_column + 1)]

        sku = str(vals[col["sku"]] or "").strip()
        if not sku:
            continue

        rfid_raw = str(vals[col["rfid_enabled"]] or "FALSE").upper().strip()
        rfid     = rfid_raw in ("TRUE", "YES", "1")

        payload = {
            "sku":            sku,
            "name":           str(vals[col["name"]] or "").strip(),
            "rfidEnabled":    rfid,
        }

        if "product_code"  in col and vals[col["product_code"]]:
            payload["erpProductCode"]  = str(vals[col["product_code"]]).strip()
        if "description"   in col and vals[col["description"]]:
            payload["description"]     = str(vals[col["description"]]).strip()
        if "brand"         in col and vals[col["brand"]]:
            payload["brand"]           = str(vals[col["brand"]]).strip()
        if "uom"           in col and vals[col["uom"]]:
            payload["unitOfMeasure"]   = str(vals[col["uom"]]).strip()
        if "weight_grams"  in col and vals[col["weight_grams"]]:
            payload["weightGrams"]     = int(vals[col["weight_grams"]])

        existing_id = client.find_product_by_sku(sku)

        if dry_run:
            action = "UPDATE" if existing_id else "CREATE"
            info(f"[DRY-RUN] {action} {sku} — {payload.get('name','')}")
            continue

        if existing_id:
            r = client.put(f"/api/products/{existing_id}", payload)
            if r.status_code == 200:
                ok(f"Updated  {sku}")
                stats["updated"] += 1
            else:
                err(f"Failed to update {sku}: {r.status_code} {r.text[:120]}")
                stats["errors"]  += 1
                continue
            product_id = existing_id
        else:
            r = client.post("/api/products", payload)
            if r.status_code == 201:
                product_id = r.json()["data"]["id"]
                ok(f"Created  {sku} → {product_id}")
                stats["created"] += 1
            else:
                err(f"Failed to create {sku}: {r.status_code} {r.text[:120]}")
                stats["errors"]  += 1
                continue

        # ── EPC tag association ───────────────────────────────────────────
        epc_raw = str(vals[col["epc_tags"]] or "") if "epc_tags" in col else ""
        epcs    = [e.strip().upper() for e in epc_raw.split(",") if e.strip()]
        for epc in epcs:
            r2 = client.session.post(
                f"{client.base_url}/api/products/{product_id}/epc",
                params={"epc": epc}, timeout=10)
            if r2.status_code in (200, 201, 409):
                ok(f"  EPC {epc} → {sku}")
                stats["epcs"] += 1
            else:
                warn(f"  EPC {epc} failed: {r2.status_code} {r2.text[:80]}")

        time.sleep(0.05)  # avoid rate limiting

    print()
    print(f"  Products  created={stats['created']}  updated={stats['updated']}  "
          f"errors={stats['errors']}  epcs_linked={stats['epcs']}")
    return stats


# =============================================================================
# INVENTORY (SOH EXPECTED QTY) UPLOAD
# =============================================================================

def upload_inventory(client: StoreLenseClient, xls_path: str,
                     dry_run: bool = False) -> dict:
    step(f"Uploading ERP SOH Expected Inventory from {xls_path}")

    wb = openpyxl.load_workbook(xls_path, data_only=True)
    ws = wb["ERP_SOH_Expected"]

    headers = [str(ws.cell(1, c).value).strip().lower()
               for c in range(1, ws.max_column + 1)]

    required = {"store_code", "sku", "qty_expected"}
    missing  = required - set(headers)
    if missing:
        sys.exit(f"{RED}Missing required columns: {missing}{RESET}")

    col = {h: i for i, h in enumerate(headers)}

    # Cache store and zone lookups
    info("Loading store and zone data...")
    stores     = client.get_stores()
    zone_cache = {}  # store_id → {zone_code: zone_id}

    stats = {"upserted": 0, "skipped": 0, "errors": 0}

    for row_num in range(2, ws.max_row + 1):
        vals = [ws.cell(row_num, c).value for c in range(1, ws.max_column + 1)]

        store_code = str(vals[col["store_code"]] or "").strip()
        sku        = str(vals[col["sku"]]        or "").strip()
        qty_raw    = vals[col["qty_expected"]]

        if not store_code or not sku or qty_raw is None:
            continue

        try:
            qty_expected = int(qty_raw)
        except (ValueError, TypeError):
            err(f"Row {row_num}: qty_expected '{qty_raw}' is not an integer — skipped")
            stats["skipped"] += 1
            continue

        if qty_expected < 0:
            err(f"Row {row_num}: qty_expected cannot be negative — skipped")
            stats["skipped"] += 1
            continue

        # Resolve store
        store_id = stores.get(store_code)
        if not store_id:
            err(f"Row {row_num}: store_code '{store_code}' not found in StoreLense — skipped")
            stats["skipped"] += 1
            continue

        # Resolve zone (optional)
        zone_code = str(vals[col["zone_code"]] or "").strip() if "zone_code" in col else ""
        zone_id   = None
        if zone_code:
            if store_id not in zone_cache:
                zone_cache[store_id] = client.get_zones(store_id)
            zone_id = zone_cache[store_id].get(zone_code)
            if not zone_id:
                warn(f"Row {row_num}: zone_code '{zone_code}' not found for store '{store_code}' — importing as store-level")

        # Resolve product
        product_id = client.find_product_by_sku(sku)
        if not product_id:
            err(f"Row {row_num}: SKU '{sku}' not found in StoreLense — skipped (upload product first)")
            stats["skipped"] += 1
            continue

        notes = str(vals[col["notes"]] or "").strip() if "notes" in col else ""
        valid_date = str(vals[col["valid_date"]] or str(date.today())).strip() if "valid_date" in col else str(date.today())

        label = f"{store_code} / {sku} / zone={zone_code or 'ALL'}"

        if dry_run:
            info(f"[DRY-RUN] UPSERT {label}  qty_expected={qty_expected}")
            continue

        # ── Call the ERP admin inventory upload endpoint ─────────────────
        # Formats the payload as a single-item inventory sync to the
        # internal inventory upload API route
        payload = {
            "storeId":         store_id,
            "productId":       product_id,
            "zoneId":          zone_id,
            "quantityExpected": qty_expected,
            "validDate":       valid_date,
            "notes":           notes,
            "source":          "xls_import",
        }

        r = client.post("/api/inventory/expected", payload)

        if r.status_code in (200, 201, 204):
            ok(f"Upserted {label}  qty={qty_expected}")
            stats["upserted"] += 1
        elif r.status_code == 404:
            # Endpoint may not be deployed — fall through to DB-direct note
            warn(f"Row {row_num}: /api/inventory/expected returned 404")
            warn("       Falling back to direct SQL — see --db-mode flag or run sql_import.sql")
            stats["errors"] += 1
        else:
            err(f"Row {row_num}: {label} failed {r.status_code}: {r.text[:120]}")
            stats["errors"] += 1

        time.sleep(0.03)

    print()
    print(f"  Inventory  upserted={stats['upserted']}  skipped={stats['skipped']}  errors={stats['errors']}")
    return stats


# =============================================================================
# SQL EXPORT  (alternative when /api/inventory/expected is not available)
# =============================================================================

def export_sql(client: StoreLenseClient, xls_path: str, output_path: str):
    """
    Generate a SQL script that directly upserts into inventory.inventory_state.
    Run this on the server:  psql -h localhost -U storelense_app -d storelense -f import.sql
    """
    step(f"Generating SQL import from {xls_path} → {output_path}")

    wb = openpyxl.load_workbook(xls_path, data_only=True)
    ws = wb["ERP_SOH_Expected"]
    headers = [str(ws.cell(1, c).value).strip().lower()
               for c in range(1, ws.max_column + 1)]
    col = {h: i for i, h in enumerate(headers)}

    stores     = client.get_stores()
    zone_cache = {}
    lines      = []

    lines.append("-- StoreLense ERP SOH Expected Inventory Import")
    lines.append(f"-- Generated: {datetime.utcnow().isoformat()}Z")
    lines.append("-- Run as: psql -h localhost -U storelense_app -d storelense -f import.sql")
    lines.append("")
    lines.append("BEGIN;")
    lines.append("")

    for row_num in range(2, ws.max_row + 1):
        vals       = [ws.cell(row_num, c).value for c in range(1, ws.max_column + 1)]
        store_code = str(vals[col["store_code"]] or "").strip()
        sku        = str(vals[col["sku"]]        or "").strip()
        qty_raw    = vals[col["qty_expected"]]

        if not store_code or not sku or qty_raw is None:
            continue

        try:
            qty_expected = int(qty_raw)
        except (ValueError, TypeError):
            lines.append(f"-- SKIP row {row_num}: invalid qty '{qty_raw}' for {sku}")
            continue

        store_id   = stores.get(store_code)
        if not store_id:
            lines.append(f"-- SKIP row {row_num}: store '{store_code}' not found")
            continue

        zone_code = str(vals[col["zone_code"]] or "").strip() if "zone_code" in col else ""
        zone_id   = None
        if zone_code:
            if store_id not in zone_cache:
                zone_cache[store_id] = client.get_zones(store_id)
            zone_id = zone_cache[store_id].get(zone_code)

        product_id = client.find_product_by_sku(sku)
        if not product_id:
            lines.append(f"-- SKIP row {row_num}: SKU '{sku}' not found — upload product first")
            continue

        zone_sql  = f"'{zone_id}'" if zone_id else "NULL"
        accuracy  = "ROUND(100.0 * quantity_on_hand / NULLIF(quantity_expected, 0), 2)"

        lines.append(f"-- {store_code} / {sku} / zone={zone_code or 'ALL'}  qty={qty_expected}")
        lines.append(
            f"INSERT INTO inventory.inventory_state "
            f"(store_id, product_id, zone_id, quantity_expected, quantity_on_hand, updated_at) "
            f"VALUES ('{store_id}', '{product_id}', {zone_sql}, {qty_expected}, 0, now()) "
            f"ON CONFLICT (store_id, product_id, zone_id) DO UPDATE "
            f"  SET quantity_expected = EXCLUDED.quantity_expected, "
            f"      accuracy_pct = {accuracy}, "
            f"      updated_at = now();"
        )
        lines.append("")

    lines.append("COMMIT;")
    lines.append("")
    lines.append("-- Verify: SELECT s.store_code, p.sku, i.quantity_expected, i.quantity_on_hand, i.accuracy_pct")
    lines.append("--         FROM inventory.inventory_state i")
    lines.append("--         JOIN stores.stores s ON s.id = i.store_id")
    lines.append("--         JOIN products.products p ON p.id = i.product_id")
    lines.append("--         ORDER BY s.store_code, p.sku;")

    with open(output_path, "w") as f:
        f.write("\n".join(lines))

    ok(f"SQL script saved → {output_path}")
    info(f"Run on server: psql -h localhost -U storelense_app -d storelense -f {output_path}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="StoreLense ERP Data Upload Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate XLS templates
  python upload_erp_data.py --create-templates

  # Upload products
  python upload_erp_data.py --products products.xlsx

  # Upload ERP expected inventory (SOH comparison data)
  python upload_erp_data.py --inventory inventory.xlsx

  # Upload both
  python upload_erp_data.py --products products.xlsx --inventory inventory.xlsx

  # Dry run (validate without writing)
  python upload_erp_data.py --products products.xlsx --dry-run

  # Generate SQL for direct DB import (no API required)
  python upload_erp_data.py --inventory inventory.xlsx --export-sql import.sql
        """)

    parser.add_argument("--url",       default=os.getenv("STORELENSE_URL", "http://localhost:8080"),
                        help="StoreLense API Gateway URL (default: http://localhost:8080)")
    parser.add_argument("--username",  default=os.getenv("STORELENSE_USER", "admin"))
    parser.add_argument("--password",  default=os.getenv("STORELENSE_PASS", "Admin@StoreLense1"))

    parser.add_argument("--create-templates", action="store_true",
                        help="Generate sample XLS template files and exit")
    parser.add_argument("--products",   metavar="FILE",
                        help="Path to product list XLS file")
    parser.add_argument("--inventory",  metavar="FILE",
                        help="Path to ERP expected inventory XLS file")
    parser.add_argument("--export-sql", metavar="OUTPUT",
                        help="Export a SQL script instead of calling API (use for direct DB import)")
    parser.add_argument("--dry-run",    action="store_true",
                        help="Validate files and print actions without writing anything")

    args = parser.parse_args()

    print(f"\n{BOLD}{CYAN}StoreLense ERP Data Upload Tool{RESET}")
    print(f"  Gateway : {args.url}")

    if args.create_templates:
        step("Creating XLS templates")
        create_product_template("product_list_template.xlsx")
        create_inventory_template("erp_soh_inventory_template.xlsx")
        print(f"\n{GREEN}Templates created.{RESET} Fill in your data and re-run with --products / --inventory.\n")
        sys.exit(0)

    if not args.products and not args.inventory:
        parser.print_help()
        sys.exit(0)

    if args.dry_run:
        warn("DRY-RUN mode — no data will be written")

    client = StoreLenseClient(args.url, args.username, args.password)

    if args.products:
        if not os.path.exists(args.products):
            sys.exit(f"{RED}File not found: {args.products}{RESET}")
        upload_products(client, args.products, dry_run=args.dry_run)

    if args.inventory:
        if not os.path.exists(args.inventory):
            sys.exit(f"{RED}File not found: {args.inventory}{RESET}")

        if args.export_sql:
            export_sql(client, args.inventory, args.export_sql)
        else:
            upload_inventory(client, args.inventory, dry_run=args.dry_run)

    print(f"\n{GREEN}{BOLD}Done.{RESET}\n")


if __name__ == "__main__":
    main()
