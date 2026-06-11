#!/usr/bin/env python3
"""
Convert any retail EPC export (Lenskart / optical format) → StoreLense upload CSV.

Input columns (flexible — handles missing columns gracefully):
  EPC | Description | Size | Info (brand) | Image

Output: storelense_products.xlsx  ready for seed_from_xls.py

Usage:
  python3 tools/convert_to_storelense.py --input your_file.xlsx
  python3 tools/convert_to_storelense.py --input your_file.xlsx --dept EYEWEAR --store-code LK001
"""

import argparse
import csv
import sys
import re
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Run: pip3 install openpyxl"); sys.exit(1)

# ── Column detection ──────────────────────────────────────────────────────────

EPC_ALIASES   = ['epc', 'article', 'article no', 'style', 'sku', 'item code',
                 'product code', 'barcode', 'rfid', 'tag']
DESC_ALIASES  = ['description', 'desc', 'product name', 'item name', 'name', 'product']
SIZE_ALIASES  = ['size', 'sz', 'item size']
BRAND_ALIASES = ['info', 'brand', 'brand name', 'label', 'make', 'manufacturer']
IMG_ALIASES   = ['image', 'img', 'image url', 'photo', 'picture', 'url']
DEPT_ALIASES  = ['dept', 'department', 'category', 'class', 'item class']
EAN_ALIASES   = ['ean', 'ean13', 'upc', 'item barcode', 'gtin', 'barcode no']

def _find(row_keys, aliases):
    row_lower = {k.lower().strip(): k for k in row_keys}
    for a in aliases:
        if a in row_lower:
            return row_lower[a]
    return None

def _val(row, aliases, default=''):
    key = _find(row.keys(), aliases)
    if key:
        v = row.get(key)
        if v is not None:
            return str(v).strip()
    return default

# ── EAN generation ────────────────────────────────────────────────────────────

def _ean13_check(digits12):
    """Compute EAN-13 check digit from first 12 digits."""
    s = [int(d) for d in digits12]
    total = sum(s[i] * (1 if i % 2 == 0 else 3) for i in range(12))
    return str((10 - (total % 10)) % 10)

def make_ean(seq, prefix='890123'):
    """Generate a valid EAN-13 from a sequence number."""
    body = f"{prefix}{seq:06d}"[:12].ljust(12, '0')
    return body + _ean13_check(body)

# ── Extract brand from Info column (handles "Lenskart -Air Switch" → "Lenskart Air") ──

def clean_brand(raw):
    if not raw:
        return 'HouseLabel'
    # Remove special chars around brand separators
    cleaned = re.sub(r'\s*[-–—]\s*', ' ', raw).strip()
    # Capitalise words
    return ' '.join(w.capitalize() for w in cleaned.split())

# ── Main conversion ───────────────────────────────────────────────────────────

def convert(input_path, dept_override, store_code, output_path):
    print(f"\n  Reading: {input_path}")

    # Load — support xlsx and csv
    ext = os.path.splitext(input_path)[1].lower()
    rows = []

    if ext in ('.xlsx', '.xls'):
        wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
        ws = wb.active
        iter_rows = ws.iter_rows(values_only=True)
        raw_hdr = next(iter_rows)
        headers = [str(h).strip() if h else f'col{i}' for i, h in enumerate(raw_hdr)]
        for r in iter_rows:
            if any(v is not None for v in r):
                rows.append(dict(zip(headers, r)))
        wb.close()
    elif ext == '.csv':
        with open(input_path, newline='', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
    else:
        print(f"Unsupported format: {ext}"); sys.exit(1)

    print(f"  Rows loaded: {len(rows)}")
    if not rows:
        print("  ERROR: no rows found"); sys.exit(1)

    print(f"  Columns found: {list(rows[0].keys())}")

    # ── Group rows by EPC/style ───────────────────────────────────────────────
    # Each unique EPC = one product in StoreLense
    products = {}
    ean_counter = 1

    for row in rows:
        epc_val  = _val(row, EPC_ALIASES)
        if not epc_val:
            continue

        style = epc_val.strip()

        if style not in products:
            desc  = _val(row, DESC_ALIASES, style)
            size  = _val(row, SIZE_ALIASES, 'One Size') or 'One Size'
            brand = clean_brand(_val(row, BRAND_ALIASES))
            dept  = dept_override or _val(row, DEPT_ALIASES, 'GENERAL')
            image = _val(row, IMG_ALIASES)

            # Use existing EAN if present, otherwise generate one
            existing_ean = _val(row, EAN_ALIASES)
            if existing_ean and existing_ean.isdigit() and len(existing_ean) >= 12:
                ean = existing_ean[:13].zfill(13)
            else:
                ean = make_ean(ean_counter)
                ean_counter += 1

            products[style] = {
                'style':       style,
                'description': desc,
                'brand':       brand,
                'dept':        dept,
                'size':        size,
                'barcode':     ean,
                'image':       image,
                'epcs':        [],
                'expected':    0,
                'scan':        0,
            }

        # The EPC value itself is the RFID tag for this physical item
        products[style]['epcs'].append(style)
        products[style]['expected'] += 1
        products[style]['scan']     += 1

    print(f"  Unique products (styles): {len(products)}")

    # ── Write output XLSX ─────────────────────────────────────────────────────
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Products"

    HEADERS = ['dept', 'style', 'description', 'brand', 'barcode', 'epc',
               'size', 'cost', 'price', 'expected', 'scan', 'back stock', 'sales floor']

    # Header row
    ws_out.append(HEADERS)
    hdr_fill = PatternFill("solid", fgColor="0F766E")
    for cell in ws_out[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows — one row per EPC tag (each physical item)
    for p in products.values():
        # If the product has multiple EPCs (multiple units), write one row per EPC
        epcs = p['epcs'] if p['epcs'] else [p['style']]
        for epc in epcs:
            ws_out.append([
                p['dept'],
                p['style'],
                p['description'],
                p['brand'],
                p['barcode'],
                epc,             # actual RFID EPC tag
                p['size'],
                0,               # cost — fill in your actual cost
                0,               # price — fill in your actual price
                1,               # expected: 1 unit per EPC
                1,               # scan: in_store
                0,               # back stock
                1,               # sales floor
            ])

    # Auto-width
    for col in ws_out.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws_out.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb_out.save(output_path)
    print(f"\n  ✔ Output written: {output_path}")
    print(f"  ✔ {len(products)} products, {sum(len(p['epcs']) for p in products.values())} EPC tags")
    print(f"\n  Next step — upload to server:")
    print(f"  scp {output_path} root@SERVER_IP:/opt/")
    print(f"  python3 /opt/StoreLense-new/tools/seed_from_xls.py \\")
    print(f"    --sql --dir /opt/ --store-code {store_code} \\")
    print(f"    --pg-container storelense-postgres-1 \\")
    print(f"    --url http://localhost:8080 --username admin --password Admin@StoreLense1")


if __name__ == '__main__':
    ap = argparse.ArgumentParser(description='Convert retail EPC export → StoreLense upload format')
    ap.add_argument('--input',      required=True,             help='Input .xlsx or .csv file')
    ap.add_argument('--output',     default='storelense_products.xlsx', help='Output .xlsx filename')
    ap.add_argument('--dept',       default='EYEWEAR',         help='Department / category override')
    ap.add_argument('--store-code', default='STORE001',        help='Your store code (e.g. LK001)')
    args = ap.parse_args()
    convert(args.input, args.dept, args.store_code, args.output)
